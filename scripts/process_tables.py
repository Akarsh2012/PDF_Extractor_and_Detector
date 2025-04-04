import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment

def is_table_line(line):
    """Detect if a line is part of a table based on presence of numbers and formatting."""
    return bool(re.search(r'\d', line)) and len(line.strip()) > 0

def clean_cell(cell):
    """Clean unwanted characters and strip spaces."""
    return re.sub(r"[\x00-\x1F\x7F]", "", cell).strip()

def process_text_to_table(txt_file, output_excel):
    with open(txt_file, "r", encoding="utf-8") as file:
        lines = file.readlines()

    tables = []
    current_table = []

    for line in lines:
        line = line.strip()
        
        if "--- End of Page" in line:  # Page separator
            if current_table:
                tables.append(current_table)
                current_table = []
            continue
        
        if is_table_line(line):
            cleaned_row = [clean_cell(cell) for cell in re.split(r'\s{2,}|\t', line)]  # Split using multiple spaces or tabs
            current_table.append(cleaned_row)

    if current_table:
        tables.append(current_table)

    # Save tables to Excel
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        for idx, table in enumerate(tables):
            df = pd.DataFrame(table)
            df.to_excel(writer, index=False, header=False, sheet_name=f'Table_{idx+1}')
    
    # Beautify the Excel file
    beautify_excel(output_excel)

    print(f"✅ Perfectly formatted tables exported to: {output_excel}")

def beautify_excel(file_path):
    """Apply styles like borders, alignment, and bold headers to enhance readability."""
    wb = load_workbook(file_path)

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        max_col = ws.max_column
        max_row = ws.max_row

        # Set column width for better visibility
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # Apply styles
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                if cell.value:  # Apply border and alignment only to filled cells
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # Make the first row bold (header styling)
        for cell in ws[1]:
            cell.font = Font(bold=True)

    wb.save(file_path)
    print("🎨 Excel formatting applied for better readability!")

if __name__ == "__main__":
    input_txt = "data/output.txt"  # Your extracted text file
    output_excel = "output/output.xlsx"
    process_text_to_table(input_txt, output_excel)
